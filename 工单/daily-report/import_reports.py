"""导入王涛的日报数据到数据库 - 通过 SQL 文件方式"""
import openpyxl
import glob
import re
import subprocess
import json
from datetime import datetime, date
from collections import defaultdict
import os

ENGINEER_NAME = '王涛'
ENGINEER_ID = 1

def safe_print(text):
    """安全打印，避免 GBK 编码错误"""
    try:
        print(text)
    except UnicodeEncodeError:
        print(repr(text))

# 重写 print 行为
_orig_print = print
import builtins
def _safe_print(*args, **kwargs):
    try:
        _orig_print(*args, **kwargs)
    except UnicodeEncodeError:
        _orig_print(*[repr(a) if isinstance(a, str) else a for a in args], **kwargs)
builtins.print = _safe_print

def esc(val):
    """SQL 字符串转义"""
    if val is None:
        return 'NULL'
    s = str(val).replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"

def esc_or_null(val):
    if val is None:
        return 'NULL'
    s = str(val).replace("\\", "\\\\").replace("'", "\\'")
    if s.strip() in ('', 'None', '/'):
        return 'NULL'
    return f"'{s}'"

# 查找所有文件
files = []
files += glob.glob(r'C:\Users\2914\Desktop\王涛*.xlsx', recursive=True)
files += glob.glob(r'C:\Users\2914\Desktop\项目\**\王涛*.xlsx', recursive=True)
files = sorted(set(f for f in files if os.path.isfile(f)))
print(f'找到 {len(files)} 个文件')

def parse_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        val = val.strip()
        for p in [r'(\d{4})\.(\d{1,2})\.(\d{1,2})', r'(\d{4})[/](\d{1,2})[/](\d{1,2})', r'(\d{4})-(\d{1,2})-(\d{1,2})']:
            m = re.match(p, val)
            if m:
                y, mo, d = m.groups()
                return f'{y}-{int(mo):02d}-{int(d):02d}'
    return None

def parse_tasks(text):
    if not text or str(text).strip() in ('/', '', '无'):
        return []
    lines = []
    for line in str(text).strip().split('\n'):
        line = line.strip()
        if not line:
            continue
        cleaned = re.sub(r'^\d+[.、．]\s*', '', line)
        if cleaned:
            lines.append(cleaned)
    return lines if lines else [str(text).strip()]

def extract_issues(text):
    if not text or str(text).strip() in ('/', '', '无'):
        return None
    t = str(text)
    m = re.search(r'遗留问题[：:]\s*(.*?)$', t, re.DOTALL)
    if m:
        issues = m.group(1).strip()
        return issues if issues and issues != '无' else None
    m = re.search(r'未完成[。.，,]\s*(.*)', t)
    if m:
        return m.group(1).strip() or None
    return t.strip() if t.strip() != '无' else None

def determine_status(progress_pct, issues_text, completion_text):
    if progress_pct >= 100:
        return 'complete'
    ct = str(completion_text or '')
    if '遗留问题' in ct and '无' not in ct.split('遗留问题')[-1][:10]:
        return 'blocked'
    return 'ongoing'

# ====== 解析所有文件 ======
project_data = {}
all_reports = []

for f in files:
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f'无法打开: {os.path.basename(f)} -> {e}')
        continue

    try:
        cell11 = str(ws.cell(1, 1).value or '')
        # 尝试匹配 (原厂工单-XXXXX) 或 （原厂工单-XXXXX）
        m = re.search(r'[（(]原厂工单[-=]?\s*(\w+)[）)]', cell11)
        order_no = m.group(1) if m else None

        # 项目名：去掉工单号部分
        project_name = re.sub(r'\s*[（(]原厂工单[-=]?\s*\w+[）)]', '', cell11).strip()
        # 去掉末尾的 +A1:H13 之类
        project_name = re.sub(r'\+\w+:\w+$', '', project_name).strip()
        if not project_name:
            # 从文件名提取
            project_name = os.path.basename(f).replace('.xlsx', '').strip()

        customer = str(ws.cell(2, 4).value or '').strip()
        manufacturer = str(ws.cell(4, 2).value or '').strip()
        agent = str(ws.cell(3, 7).value or '').strip()
        contact_full = str(ws.cell(3, 4).value or '').strip()
        contact_name, contact_phone = contact_full, ''
        if '/' in contact_full:
            parts = contact_full.split('/')
            contact_name = parts[0].strip()
            contact_phone = '/'.join(parts[1:]).strip() if len(parts) > 1 else ''
        address = str(ws.cell(8, 2).value or '').strip()
        tech_lead_full = str(ws.cell(5, 2).value or '').strip()
        tech_lead = tech_lead_full.split('/')[0].strip() if tech_lead_full else ''
        service_manager = str(ws.cell(4, 7).value or '').strip()

        key = order_no or project_name
        if key not in project_data:
            project_data[key] = {
                'name': project_name, 'order_no': order_no, 'customer': customer,
                'manufacturer': manufacturer, 'agent': agent,
                'contact_name': contact_name, 'contact_phone': contact_phone,
                'address': address, 'tech_lead': tech_lead, 'service_manager': service_manager
            }

        # 解析日报行
        daily_count = 0
        for r in range(13, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            plan_title = ws.cell(r, 2).value
            if not date_val or not plan_title:
                continue
            report_date = parse_date(date_val)
            if not report_date:
                continue
            pt = str(plan_title).strip()
            if not pt or pt == '工作计划':
                continue

            work_content = str(ws.cell(r, 3).value or '').strip()
            tasks = parse_tasks(work_content)
            if not tasks:
                tasks = [pt]

            progress_val = ws.cell(r, 8).value
            progress_pct = 0
            if progress_val is not None:
                try:
                    p = float(progress_val)
                    progress_pct = int(p * 100) if p <= 1 else int(p)
                except:
                    pass

            completion_text = str(ws.cell(r, 6).value or '').strip()
            issues = extract_issues(completion_text)
            status = determine_status(progress_pct, completion_text, completion_text)

            all_reports.append((key, report_date, pt, tasks, progress_pct, status, issues))
            daily_count += 1

        fname = os.path.basename(f).replace('\xa0', ' ')
        print(f'  {fname:50s} {daily_count} 条日报')
    except Exception as e:
        fname = os.path.basename(f).replace('\xa0', ' ')
        print(f'  解析失败 {fname}: {e[:80] if isinstance(e, str) else str(e)[:80]}')

print(f'\n解析完成: {len(project_data)} 个项目, {len(all_reports)} 条日报')

# ====== 生成 SQL 文件 ======
sql_lines = ['SET NAMES utf8mb4;']
sql_lines.append('START TRANSACTION;')

# 1. 插入项目
project_id_map = {}
for key, pinfo in project_data.items():
    name = pinfo['name']
    sql_lines.append(
        f"INSERT IGNORE INTO projects (name, order_no, customer, contact_name, contact_phone, "
        f"install_address, manufacturer, agent, tech_lead, service_manager, created_by) VALUES "
        f"({esc(name)}, {esc_or_null(pinfo['order_no'])}, {esc(pinfo['customer'])}, "
        f"{esc_or_null(pinfo['contact_name'])}, {esc_or_null(pinfo['contact_phone'])}, "
        f"{esc_or_null(pinfo['address'])}, {esc_or_null(pinfo['manufacturer'])}, "
        f"{esc_or_null(pinfo['agent'])}, {esc_or_null(pinfo['tech_lead'])}, "
        f"{esc_or_null(pinfo['service_manager'])}, {ENGINEER_ID});"
    )

# 2. 确保工程师关联
for key in project_data:
    if project_data[key]['order_no']:
        cond = f"order_no = {esc(project_data[key]['order_no'])}"
    else:
        cond = f"name = {esc(project_data[key]['name'])}"
    sql_lines.append(
        f"INSERT IGNORE INTO engineer_projects (engineer_id, project_id, impl_days) "
        f"SELECT {ENGINEER_ID}, id, 0 FROM projects WHERE {cond} LIMIT 1;"
    )

# 3. 按项目+日期排序，插入日报
reports_by_project = defaultdict(list)
for key, rdate, pt, tasks, prog, st, issues in all_reports:
    reports_by_project[key].append((rdate, pt, tasks, prog, st, issues))

for key, reports in reports_by_project.items():
    reports.sort(key=lambda x: x[0])
    # 获取 project_id
    pinfo = project_data[key]
    if pinfo['order_no']:
        cond = f"order_no = {esc(pinfo['order_no'])}"
    else:
        cond = f"name = {esc(pinfo['name'])}"

    impl_day = 0
    for rdate, pt, tasks, prog, st, issues in reports:
        impl_day += 1
        tasks_json = json.dumps(tasks, ensure_ascii=False)
        next_plan = pt if st == 'ongoing' else None
        sql_lines.append(
            f"INSERT IGNORE INTO daily_reports "
            f"(engineer_id, project_id, report_date, impl_day, plan_title, tasks, progress, status, issues, next_plan) "
            f"SELECT {ENGINEER_ID}, id, '{rdate}', {impl_day}, "
            f"{esc(pt)}, {esc(tasks_json)}, {prog}, '{st}', "
            f"{esc_or_null(issues)}, {esc_or_null(next_plan)} "
            f"FROM projects WHERE {cond} LIMIT 1;"
        )

# 4. 统一更新 impl_days
sql_lines.append(
        f"UPDATE engineer_projects ep SET ep.impl_days = ("
        f"SELECT COUNT(*) FROM daily_reports dr "
        f"WHERE dr.engineer_id = ep.engineer_id AND dr.project_id = ep.project_id"
        f") WHERE ep.engineer_id = {ENGINEER_ID};"
    )

sql_lines.append('COMMIT;')

# 写入 SQL 文件
sql_path = r'C:\Users\2914\Desktop\工单\daily-report\import_reports.sql'
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_lines))

print(f'\nSQL 文件已生成: {sql_path}')
print(f'共 {len(sql_lines)} 行 SQL')
print(f'项目: {len(project_data)}, 日报: {len(all_reports)}')

# ====== 执行 SQL ======
print('\n===== 开始导入数据库 =====')
result = subprocess.run(
    r'kubectl exec -i mysql-0 -n daily-report -- mysql -u reporter -preporter123 daily_report --default-character-set=utf8mb4 < "C:\Users\2914\Desktop\工单\daily-report\import_reports.sql"',
    shell=True,
    capture_output=True,
    text=True,
    timeout=120
)

if result.returncode == 0:
    print('导入成功！')
    # 验证
else:
    print(f'导入错误: {result.stderr[:500]}')
    if result.stdout:
        print(f'输出: {result.stdout[:500]}')

# 验证
print('\n===== 验证数据 =====')
verify = subprocess.run(
    r'kubectl exec mysql-0 -n daily-report -- mysql -u reporter -preporter123 daily_report --default-character-set=utf8mb4 -e "SELECT p.name AS 项目, COUNT(dr.id) AS 日报数, MAX(ep.impl_days) AS 实施天数 FROM projects p LEFT JOIN daily_reports dr ON dr.project_id = p.id AND dr.engineer_id=1 LEFT JOIN engineer_projects ep ON ep.project_id=p.id AND ep.engineer_id=1 WHERE p.created_by=1 OR dr.id IS NOT NULL GROUP BY p.id ORDER BY 日报数 DESC;"',
    shell=True,
    capture_output=True,
    text=True,
    timeout=15
)
print(verify.stdout or verify.stderr[:500])
